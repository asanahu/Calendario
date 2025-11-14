from flask import Flask, render_template, request, redirect, url_for, jsonify, abort, flash, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import datetime
import boto3
from datetime import datetime, timedelta, timezone, date
from calendar import monthrange
from dotenv import load_dotenv
from functools import wraps
from pymongo import MongoClient, ASCENDING
from bson.objectid import ObjectId
from bson import ObjectId
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import time
import uuid
import json
from io import BytesIO

# Cargar variables de entorno
load_dotenv()
MONGO_URI = os.getenv("MONGO_URI")
AWS_S3_BUCKET = os.getenv("AWS_S3_BUCKET")
AWS_S3_REGION = os.getenv("AWS_S3_REGION")

# Conectar a MongoDB
client = MongoClient(MONGO_URI)
db = client["calendario"]
users_collection = db["usuarios"]
events_collection = db["eventos"]
historial_collection = db["historial_conversaciones"]

ruta_faqs = "faqs_generadas.json"

# 🔹 Sistema de caché mejorado para eventos
events_cache = {}
CACHE_DURATION = 60  # 1 minuto en segundos (más corto para datos críticos)
last_data_modification = time.time()  # Timestamp de la última modificación de datos

DEMO_ADMIN_USERS = {"admin"}


def is_demo_admin_user():
    """Return True when the current session belongs to the demo admin."""
    try:
        return current_user.is_authenticated and current_user.usuario in DEMO_ADMIN_USERS
    except Exception:
        return False


def build_anonymized_label_map(visible_only=True, include_fullname_map=False):
    """Generate deterministic masked labels for users.

    Args:
        visible_only: cuando es True solo considera usuarios visibles en calendario.
        include_fullname_map: devuelve un segundo diccionario {nombre completo: alias}.
    """
    query = {"visible_calendario": {"$ne": False}} if visible_only else {}
    usuarios = list(
        users_collection.find(
            query,
            {"_id": 1, "nombre": 1, "apellidos": 1}
        ).sort([('nombre', ASCENDING), ('apellidos', ASCENDING)])
    )
    id_map = {}
    fullname_map = {}
    for idx, usuario in enumerate(usuarios, start=1):
        label = f"Trabajador {idx:02d}"
        user_id = str(usuario["_id"])
        id_map[user_id] = label
        fullname = f"{usuario.get('nombre', '')} {usuario.get('apellidos', '')}".strip()
        if fullname:
            fullname_map[fullname] = label
    if include_fullname_map:
        return id_map, fullname_map
    return id_map


def get_display_name_for_user(usuario, anonymized_labels=None):
    """Return the name to show in UI, masking it when required."""
    real_name = f"{usuario.get('nombre', '')} {usuario.get('apellidos', '')}".strip()
    if anonymized_labels:
        label = anonymized_labels.get(str(usuario["_id"]))
        if label:
            return label
    return real_name


def resolve_persona_nombre(persona_param):
    """Map persona filter values to a real full name when possible."""
    if not persona_param or persona_param == 'todos':
        return None
    if ObjectId.is_valid(persona_param):
        usuario = users_collection.find_one({"_id": ObjectId(persona_param)}, {"nombre": 1, "apellidos": 1})
        if usuario:
            return f"{usuario.get('nombre', '')} {usuario.get('apellidos', '')}".strip()
    return persona_param


def get_cache_key(estados_filtro, persona_filtro, rol_filtro, busqueda):
    """Genera una clave única para el caché basada en los filtros"""
    return f"{sorted(estados_filtro)}_{persona_filtro}_{rol_filtro}_{busqueda}"

def is_cache_valid(cache_entry):
    """Verifica si una entrada del caché sigue siendo válida"""
    current_time = time.time()
    # El caché es válido si:
    # 1. No ha expirado por tiempo
    # 2. No se han modificado datos desde que se creó
    return (current_time - cache_entry['timestamp'] < CACHE_DURATION and 
            cache_entry['timestamp'] > last_data_modification)

def invalidate_cache():
    """Invalida todo el caché cuando se modifican datos críticos"""
    global last_data_modification
    last_data_modification = time.time()
    events_cache.clear()
    print("🗑️ Caché invalidado por modificación de datos")

# Lista de festivos utilizados en la aplicación
FESTIVOS = {
    # Festivos 2025
    "2025-01-01", "2025-01-06", "2025-01-29", "2025-03-05", "2025-03-28", 
    "2025-03-29", "2025-04-23", "2025-05-01", "2025-08-15", "2025-10-12",
    "2025-11-01", "2025-12-06", "2025-12-08", "2025-12-25", "2026-01-01",
    "2025-04-17", "2025-04-18", "2025-12-24", "2025-12-31",
    # Festivos 2026
    "2026-01-01", "2026-01-06", "2026-01-29", "2026-03-05", "2026-04-23", 
    "2026-05-01", "2026-08-15", "2026-10-12", "2026-11-02", "2026-12-07", 
    "2026-12-08", "2026-12-25", "2026-12-24", "2026-12-31",
    # Festivos 2027 (para futuro)
    "2027-01-01", "2027-01-06", "2027-01-29", "2027-03-05", "2027-04-23",
    "2027-05-01", "2027-08-16", "2027-10-12", "2027-11-01", "2027-12-06", 
    "2027-12-08", "2027-12-25", "2027-12-24", "2027-12-31"
}

FESTIVOS_LIST = sorted(FESTIVOS)
FESTIVOS_DATES = {datetime.strptime(fecha, "%Y-%m-%d").date() for fecha in FESTIVOS_LIST}


def es_dia_habil(fecha):
    """Devuelve True si la fecha no cae en fin de semana ni en la lista de festivos."""
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha, "%Y-%m-%d")
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    return fecha.weekday() < 5 and fecha not in FESTIVOS_DATES


def contar_dias_habiles_en_rango(fecha_inicio, fecha_fin):
    """Cuenta los días hábiles dentro de un rango cerrado, excluyendo festivos y fines de semana."""
    if isinstance(fecha_inicio, str):
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    if isinstance(fecha_fin, str):
        fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
    fecha_actual = fecha_inicio.date()
    fecha_fin_date = fecha_fin.date()
    dias = 0
    while fecha_actual <= fecha_fin_date:
        if es_dia_habil(fecha_actual):
            dias += 1
        fecha_actual += timedelta(days=1)
    return dias


def normalizar_fecha_str(fecha_valor):
    if isinstance(fecha_valor, datetime):
        return fecha_valor.strftime("%Y-%m-%d")
    if isinstance(fecha_valor, str):
        return fecha_valor
    if hasattr(fecha_valor, "strftime"):
        return fecha_valor.strftime("%Y-%m-%d")
    return str(fecha_valor)


def limpiar_vacaciones_duplicadas(trabajador):
    """Elimina duplicados exactos (mismo día) para un trabajador manteniendo un único registro."""
    cursor = events_collection.find(
        {"trabajador": trabajador, "tipo": "Vacaciones"},
        sort=[("fecha_inicio", ASCENDING), ("_id", ASCENDING)]
    )
    vistos = set()
    ids_a_borrar = []
    for vacacion in cursor:
        inicio = normalizar_fecha_str(vacacion.get("fecha_inicio"))
        fin = normalizar_fecha_str(vacacion.get("fecha_fin"))
        clave = (inicio, fin)
        if clave in vistos:
            ids_a_borrar.append(vacacion["_id"])
        else:
            vistos.add(clave)
    if ids_a_borrar:
        events_collection.delete_many({"_id": {"$in": ids_a_borrar}})
        invalidate_cache()  # Invalidar caché al limpiar duplicados


def filtrar_vacaciones_unicas(vacaciones):
    """Devuelve sólo un registro por día para evitar duplicados en memoria."""
    resultado = []
    vistos = set()
    for vacacion in vacaciones:
        inicio = normalizar_fecha_str(vacacion.get("fecha_inicio"))
        fin = normalizar_fecha_str(vacacion.get("fecha_fin"))
        clave = (inicio, fin)
        if clave in vistos:
            continue
        vistos.add(clave)
        resultado.append(vacacion)
    return resultado


VACATION_CYCLES = [
    {"label": "2025", "inicio": date(2025, 1, 15), "fin": date(2026, 1, 15)},
    {"label": "2026", "inicio": date(2026, 1, 15), "fin": date(2027, 1, 15)},
    {"label": "2027", "inicio": date(2027, 1, 15), "fin": date(2028, 1, 15)}
]


def obtener_ciclo_label(fecha):
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    for idx, ciclo in enumerate(VACATION_CYCLES):
        inicio = ciclo["inicio"]
        fin = ciclo["fin"]

        if fecha == inicio and idx > 0:
            return VACATION_CYCLES[idx - 1]["label"]

        if inicio <= fecha <= fin:
            return ciclo["label"]

    if fecha < VACATION_CYCLES[0]["inicio"]:
        return VACATION_CYCLES[0]["label"]
    return VACATION_CYCLES[-1]["label"]


def dividir_grupo_por_ciclo(grupo):
    if not grupo:
        return []
    subgrupos = []
    ordenados = sorted(grupo, key=lambda v: v["fecha_inicio"])
    ciclo_actual = obtener_ciclo_label(ordenados[0]["fecha_inicio"])
    acumulado = []
    for vacacion in ordenados:
        ciclo_vacacion = obtener_ciclo_label(vacacion["fecha_inicio"])
        if ciclo_vacacion != ciclo_actual:
            if acumulado:
                subgrupos.append((ciclo_actual, acumulado))
            acumulado = []
            ciclo_actual = ciclo_vacacion
        acumulado.append(vacacion)
    if acumulado:
        subgrupos.append((ciclo_actual, acumulado))
    return subgrupos


def preparar_vacaciones_para_template(grupos_vacaciones):
    ciclos_preparados = []
    ciclos_index = {}
    for ciclo in VACATION_CYCLES:
        entrada = {
            "label": ciclo["label"],
            "inicio": ciclo["inicio"],
            "fin": ciclo["fin"],
            "fin_inclusivo": ciclo["fin"],
            "grupos": [],
            "total_dias_habiles": 0
        }
        ciclos_preparados.append(entrada)
        ciclos_index[ciclo["label"]] = entrada

    total_dias = 0
    total_grupos = 0

    for grupo in grupos_vacaciones:
        for ciclo_label, subgrupo in dividir_grupo_por_ciclo(grupo):
            if not subgrupo:
                continue
            inicio = subgrupo[0]["fecha_inicio"]
            fin = subgrupo[-1]["fecha_fin"]
            dias_habiles = contar_dias_habiles_en_rango(inicio, fin)
            informacion = {
                "inicio": inicio,
                "fin": fin,
                "dias_habiles": dias_habiles,
                "vacaciones": subgrupo
            }
            ciclos_index[ciclo_label]["grupos"].append(informacion)
            ciclos_index[ciclo_label]["total_dias_habiles"] += dias_habiles
            total_dias += dias_habiles
            total_grupos += 1

    ciclos_con_datos = [ciclo for ciclo in ciclos_preparados if ciclo["grupos"]]
    return ciclos_con_datos, total_dias, total_grupos

# Inicializar el cliente de boto3
s3_client = boto3.client(
    's3',
    region_name=AWS_S3_REGION,
    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY")
)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv("SECRET_KEY")

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

class User(UserMixin):
    def __init__(self, user_data):
        self.id = str(user_data['_id'])
        self.nombre = user_data['nombre']
        self.apellidos = user_data['apellidos']
        self.usuario = user_data['usuario']
        self.puesto = user_data['puesto']


@login_manager.user_loader
def load_user(user_id):
    """ Cargar usuario desde MongoDB por ID """
    from bson import ObjectId
    try:
        user_data = users_collection.find_one({"_id": ObjectId(user_id)})
        return User(user_data) if user_data else None
    except Exception as e:
        print(f"Error al cargar usuario {user_id}: {e}")
        return None

# Decorador para restringir acceso a administradores
def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.puesto != "Administrador/a":
            abort(403)
        return func(*args, **kwargs)
    return wrapper

# Decorador para restringir acceso solo a los super administradores
def super_admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.usuario not in ["sclavero", "asanahuja"]:
            abort(403)
        return func(*args, **kwargs)
    return wrapper

@app.route('/')
def home():
    if current_user.is_authenticated:
        logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/calendar')
@login_required
def calendar_page():
    # Obtener todos los usuarios visibles
    usuarios = list(users_collection.find(
        {"visible_calendario": {"$ne": False}},
        {"nombre": 1, "apellidos": 1}
    ))

    anonymized_labels = build_anonymized_label_map() if is_demo_admin_user() else None
    usuarios_context = []
    for usuario in usuarios:
        usuario_id = str(usuario['_id'])
        display_name = get_display_name_for_user(usuario, anonymized_labels)
        usuarios_context.append({
            "id": usuario_id,
            "display_name": display_name
        })

    usuarios_context.sort(key=lambda x: x["display_name"].lower())
    return render_template('index.html', usuarios=usuarios_context)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario_input = request.form['usuario']
        password_input = request.form['password']
        user_data = users_collection.find_one({"usuario": usuario_input})

        if user_data is None or not check_password_hash(user_data.get("password", ""), password_input):
            flash("Usuario o contraseña incorrectos", "error")
            return redirect(url_for('login'))
        
        user = User(user_data)
        user.id = str(user_data["_id"])  # Asegurar que el ID sea string
        login_user(user)
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

"""
@app.route('/calendar')
@login_required
def calendar_page():
    return render_template('index.html')
"""
@app.route('/ai-assistant')
@login_required
def ai_assistant():
    return render_template('ai_assistant.html')


def get_color_by_puesto(puesto):
    clases = {
        "TS": "ts-event",
        "ADM": "adm-event",
        "Administrador/a": "administrador-event",
    }
    return clases.get(puesto, "default-event")  # 🔹 Si el puesto no está en la lista, usa una clase por defecto


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    usuarios = list(users_collection.find().sort("nombre"))
    demo_mode = is_demo_admin_user()
    anonymized_labels = build_anonymized_label_map(visible_only=False) if demo_mode else None
    for usuario in usuarios:
        usuario['display_name'] = get_display_name_for_user(usuario, anonymized_labels)
    return render_template('admin_users.html', usuarios=usuarios, hide_real_names=demo_mode)

@app.route('/admin/add_user', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if request.method == 'POST':
        password = request.form.get('password')
        visible_str = request.form.get('visible_calendario', 'false')

        # Convertir la cadena recibida en un booleano de forma segura
        visible = visible_str.lower() == "true"
        
        user_data = {
            "nombre": request.form.get('nombre'),
            "apellidos": request.form.get('apellidos'),
            "usuario": request.form.get('usuario'),
            "puesto": request.form.get('puesto'),
            "password": generate_password_hash(password),
            "visible_calendario": visible
        }
        if users_collection.find_one({"usuario": user_data["usuario"]}):
            return "El usuario ya existe", 400
        users_collection.insert_one(user_data)
        invalidate_cache()  # Invalidar caché al añadir usuario
        return redirect('/admin/users')
    return render_template('add_user.html')

@app.route('/admin/edit_user/<user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    usuario = users_collection.find_one({"_id": ObjectId(user_id)})

    if not usuario:
        abort(404)

    usuario.setdefault('visible_calendario', False)

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        apellidos = request.form.get('apellidos', '').strip()
        usuario_login = request.form.get('usuario', '').strip()
        puesto = request.form.get('puesto', '').strip()
        visible_str = request.form.get('visible_calendario', 'false')
        visible = visible_str.lower() == 'true'

        usuario_actualizado = dict(usuario)
        usuario_actualizado.update({
            'nombre': nombre,
            'apellidos': apellidos,
            'usuario': usuario_login,
            'puesto': puesto,
            'visible_calendario': visible
        })

        if not all([nombre, apellidos, usuario_login, puesto]):
            flash('Todos los campos son obligatorios', 'error')
            return render_template('edit_user.html', usuario=usuario_actualizado), 400

        existing_user = users_collection.find_one({"usuario": usuario_login, "_id": {"$ne": ObjectId(user_id)}})
        if existing_user:
            flash('El nombre de usuario ya existe', 'error')
            return render_template('edit_user.html', usuario=usuario_actualizado), 400

        nombre_anterior = f"{usuario['nombre']} {usuario['apellidos']}".strip()
        nuevo_nombre_completo = f"{nombre} {apellidos}".strip()

        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {
                'nombre': nombre,
                'apellidos': apellidos,
                'usuario': usuario_login,
                'puesto': puesto,
                'visible_calendario': visible
            }}
        )

        if nombre_anterior != nuevo_nombre_completo:
            events_collection.update_many(
                {"trabajador": nombre_anterior},
                {"$set": {"trabajador": nuevo_nombre_completo}}
            )

        invalidate_cache()  # Invalidar caché al modificar usuario
        flash('Usuario actualizado correctamente', 'success')
        return redirect(url_for('admin_users'))

    return render_template('edit_user.html', usuario=usuario)

@app.route('/admin/delete_user/<user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    usuario = users_collection.find_one({"_id": ObjectId(user_id)})

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    # 🚨 No permitir que un administrador se elimine a sí mismo
    if usuario["usuario"] == current_user.usuario:
        return jsonify({"message": "No puedes eliminarte a ti mismo"}), 403

    users_collection.delete_one({"_id": ObjectId(user_id)})
    invalidate_cache()  # Invalidar caché al eliminar usuario
    return redirect('/admin/users')

# Página para que los super administradores puedan resetear contraseñas
@app.route('/admin/reset_passwords')
@login_required
@super_admin_required
def reset_passwords():
    usuarios = list(users_collection.find().sort("nombre"))
    return render_template('reset_passwords.html', usuarios=usuarios)

# Acción para actualizar la contraseña de un usuario
@app.route('/admin/reset_password/<user_id>', methods=['POST'])
@login_required
@super_admin_required
def reset_password(user_id):
    new_password = request.form.get('new_password')
    if not new_password:
        flash("Debe indicar una nueva contraseña", "error")
        return redirect(url_for('reset_passwords'))
    hashed = generate_password_hash(new_password)
    users_collection.update_one({"_id": ObjectId(user_id)}, {"$set": {"password": hashed}})
    flash("Contraseña actualizada", "success")
    return redirect(url_for('reset_passwords'))

@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Obtener el usuario actual desde la BD
        user_data = users_collection.find_one({"_id": ObjectId(current_user.id)})
        
        # Verificar que la contraseña actual sea correcta
        if not check_password_hash(user_data.get("password", ""), current_password):
            flash("La contraseña actual es incorrecta", "error")
            return redirect(url_for('cambiar_password'))
        
        # Verificar que la nueva contraseña y su confirmación coincidan
        if new_password != confirm_password:
            flash("La nueva contraseña y la confirmación no coinciden", "error")
            return redirect(url_for('cambiar_password'))
        
        # Generar el hash de la nueva contraseña y actualizar la BD
        new_hashed_password = generate_password_hash(new_password)
        users_collection.update_one(
            {"_id": ObjectId(current_user.id)},
            {"$set": {"password": new_hashed_password}}
        )
        flash("Contraseña actualizada exitosamente", "success")
        return redirect(url_for('dashboard'))
    
    return render_template("cambiar_password.html")

def agrupar_vacaciones(vacaciones):
    """
    Agrupa las vacaciones consecutivas. Se asume que la lista 'vacaciones' está ordenada por 'fecha_inicio'.
    Cada grupo es una lista de vacaciones consecutivas, donde la fecha de inicio de una vacación
    es exactamente un día después de la fecha de fin de la vacación anterior.
    """
    grupos = []
    grupo_actual = []
    for vacacion in vacaciones:
        if not grupo_actual:
            grupo_actual.append(vacacion)
        else:
            ultimo = grupo_actual[-1]
            # Comprobamos si la vacación actual es consecutiva con respecto al grupo actual.
            if vacacion["fecha_inicio"] == ultimo["fecha_fin"] + timedelta(days=1):
                grupo_actual.append(vacacion)
            else:
                grupos.append(grupo_actual)
                grupo_actual = [vacacion]
    if grupo_actual:
        grupos.append(grupo_actual)
    return grupos

@app.route('/add-vacation', methods=['GET', 'POST'])
@login_required
def add_vacation():
    if request.method == 'POST':
        vacation_data = {
            "trabajador": f"{current_user.nombre} {current_user.apellidos}",
            "fecha_inicio": request.form.get('fecha_inicio'),
            "fecha_fin": request.form.get('fecha_fin'),
            "tipo": "Vacaciones"
        }
        # Convertir las fechas a objetos datetime para iterar día a día
        fecha_inicio_dt = datetime.strptime(vacation_data["fecha_inicio"], "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(vacation_data["fecha_fin"], "%Y-%m-%d")
        
        # Insertar un evento por cada día en el rango
        current_day = fecha_inicio_dt
        while current_day <= fecha_fin_dt:
            day_str = current_day.strftime("%Y-%m-%d")
            event = {
                "trabajador": vacation_data["trabajador"],
                "fecha_inicio": day_str,
                "fecha_fin": day_str,
                "tipo": "Vacaciones"
            }
            query = {
                "trabajador": vacation_data["trabajador"],
                "fecha_inicio": day_str,
                "fecha_fin": day_str,
                "tipo": "Vacaciones"
            }
            existentes = list(events_collection.find(query, {"_id": 1}).sort("_id", ASCENDING))
            if existentes:
                ids_sobrantes = [doc["_id"] for doc in existentes[1:]]
                if ids_sobrantes:
                    events_collection.delete_many({"_id": {"$in": ids_sobrantes}})
            else:
                events_collection.insert_one(event)
            current_day += timedelta(days=1)
        limpiar_vacaciones_duplicadas(vacation_data["trabajador"])
        invalidate_cache()  # Invalidar caché al añadir vacaciones
        return redirect('/add-vacation')
    
    # Obtener las vacaciones existentes para el usuario, ordenadas por fecha de inicio
    trabajador_actual = f"{current_user.nombre} {current_user.apellidos}"
    limpiar_vacaciones_duplicadas(trabajador_actual)
    vacaciones = list(events_collection.find({
        "trabajador": trabajador_actual,
        "tipo": "Vacaciones"
    }).sort("fecha_inicio", 1))
    vacaciones = filtrar_vacaciones_unicas(vacaciones)
    
    # Convertir los strings de fechas a objetos datetime
    for vacacion in vacaciones:
        vacacion["fecha_inicio"] = datetime.strptime(vacacion["fecha_inicio"], "%Y-%m-%d")
        vacacion["fecha_fin"] = datetime.strptime(vacacion["fecha_fin"], "%Y-%m-%d")
    
    # Agrupar las vacaciones consecutivas
    grupos_vacaciones = agrupar_vacaciones(vacaciones)
    vacaciones_por_ciclo, total_dias_habiles, total_grupos = preparar_vacaciones_para_template(grupos_vacaciones)

    return render_template(
        'add_vacation.html',
        grupos_vacaciones=grupos_vacaciones,
        vacaciones_por_ciclo=vacaciones_por_ciclo,
        total_dias_habiles=total_dias_habiles,
        total_grupos=total_grupos
    )

@app.route('/add-recurring', methods=['GET', 'POST'])
@login_required
@admin_required
def add_recurring():
    """Asignar estados de forma recurrente a los trabajadores."""
    if request.method == 'POST':
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        dias_semana = request.form.getlist('dias_semana')
        n_semanas = int(request.form.get('n_semanas', 1))

        fi_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        ff_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        dias_int = [int(d) for d in dias_semana]
        dias_generados = []
        current = fi_dt
        while current <= ff_dt:
            weeks_diff = (current - fi_dt).days // 7
            if weeks_diff % n_semanas == 0 and current.weekday() in dias_int:
                dias_generados.append(current)
            current += timedelta(days=1)

        trabajadores = list(users_collection.find())
        for trabajador in trabajadores:
            estado = request.form.get(f"tipo_{trabajador['_id']}")
            if not estado:
                continue

            nombre_completo = f"{trabajador['nombre']} {trabajador['apellidos']}"
            for dia in dias_generados:
                day_str = dia.strftime("%Y-%m-%d")
                query_filter = {
                    "trabajador": nombre_completo,
                    "fecha_inicio": day_str,
                    "fecha_fin": day_str,
                    "tipo": {"$in": ["Baja", "CADE 30", "CADE 50", "CADE Tardes", "Guardia CADE", "Refuerzo Cade", "Mail"]}
                }

                if not es_dia_habil(dia):
                    if estado == "normal":
                        events_collection.delete_many(query_filter)
                    continue

                if estado == "normal":
                    events_collection.delete_many(query_filter)
                else:
                    events_collection.delete_many(query_filter)
                    nuevo_evento = {
                        "trabajador": nombre_completo,
                        "fecha_inicio": day_str,
                        "fecha_fin": day_str,
                        "tipo": estado
                    }
                    events_collection.insert_one(nuevo_evento)

        invalidate_cache()  # Invalidar caché al asignar estados recurrentes
        return redirect('/add-recurring')

    trabajadores = list(users_collection.find({"visible_calendario": True}))
    trabajadores = sorted(trabajadores, key=lambda x: (x['nombre'].lower(), x['apellidos'].lower()))
    demo_mode = is_demo_admin_user()
    anonymized_labels = build_anonymized_label_map() if demo_mode else None
    for trabajador in trabajadores:
        trabajador['display_name'] = get_display_name_for_user(trabajador, anonymized_labels)
        full_name = f"{trabajador.get('nombre', '')} {trabajador.get('apellidos', '')}".strip()
        trabajador['full_name'] = full_name
        search_source = trabajador['display_name'] if demo_mode else full_name
        trabajador['search_value'] = search_source.lower()
    return render_template('add_recurring.html', trabajadores=trabajadores, hide_real_names=demo_mode)

@app.route('/admin/user_vacations/<user_id>')
@login_required
@admin_required
def user_vacations(user_id):
    usuario = users_collection.find_one({"_id": ObjectId(user_id)})
    if not usuario:
        abort(404)

    # Obtener las vacaciones, ordenadas por fecha de inicio (ascendente)
    nombre_completo = f"{usuario['nombre']} {usuario['apellidos']}"
    limpiar_vacaciones_duplicadas(nombre_completo)
    vacaciones = list(events_collection.find(
        {"trabajador": nombre_completo, "tipo": "Vacaciones"}
    ).sort("fecha_inicio", 1))
    vacaciones = filtrar_vacaciones_unicas(vacaciones)

    # Convertir las fechas de string a objetos datetime
    for vacacion in vacaciones:
        vacacion["fecha_inicio"] = datetime.strptime(vacacion["fecha_inicio"], "%Y-%m-%d")
        vacacion["fecha_fin"] = datetime.strptime(vacacion["fecha_fin"], "%Y-%m-%d")

    # Agrupar las vacaciones consecutivas
    grupos_vacaciones = agrupar_vacaciones(vacaciones)
    vacaciones_por_ciclo, total_dias_habiles, total_grupos = preparar_vacaciones_para_template(grupos_vacaciones)

    return render_template(
        'user_vacations.html',
        usuario=usuario,
        grupos_vacaciones=grupos_vacaciones,
        vacaciones_por_ciclo=vacaciones_por_ciclo,
        total_dias_habiles=total_dias_habiles,
        total_grupos=total_grupos
    )

@app.route('/delete-vacation/<vacation_id>', methods=['POST'])
@login_required
def delete_vacation(vacation_id):
    try:
        # 🔹 Asegurar que el ID sea un ObjectId válido
        if not ObjectId.is_valid(vacation_id):
            print("❌ ID no válido para MongoDB")
            return jsonify({"error": "ID no válido"}), 400

        query = {"_id": ObjectId(vacation_id)}
        vacacion = events_collection.find_one(query)

        if vacacion and vacacion["trabajador"] == f"{current_user.nombre} {current_user.apellidos}":
            events_collection.delete_one(query)
            print("✅ Vacación eliminada correctamente")
            return redirect('/add-vacation')
        else:
            print("⚠️ No tienes permiso para eliminar esta vacación")
            return jsonify({"error": "No autorizado"}), 403

    except Exception as e:
        print(f"⚠️ Error al eliminar la vacación: {e}")
        return jsonify({"error": "Error interno"}), 500
    
    
@app.route('/api/events', methods=['GET', 'POST'])
@login_required
def events():
    if request.method == 'POST':
        event_data = request.get_json()
        event_data["trabajador"] = f"{current_user.nombre} {current_user.apellidos}"
        events_collection.insert_one(event_data)
        # Invalidar caché cuando se añade un evento
        invalidate_cache()
        return jsonify({"message": "Evento agregado correctamente"}), 201

    # 🔹 Obtener parámetros de filtro desde la query string
    estados_filtro = request.args.getlist('estados')
    persona_filtro = request.args.get('persona', 'todos')
    rol_filtro = request.args.get('rol', 'todos')
    busqueda = request.args.get('busqueda', '').strip()
    persona_nombre = resolve_persona_nombre(persona_filtro)

    demo_admin_mode = is_demo_admin_user()
    anonymized_labels = build_anonymized_label_map() if demo_admin_mode else None

    # 🔹 Verificar caché
    cache_key = f"{get_cache_key(estados_filtro, persona_filtro, rol_filtro, busqueda)}|anon={int(demo_admin_mode)}"
    if cache_key in events_cache and is_cache_valid(events_cache[cache_key]):
        print(f"🚀 Sirviendo desde caché: {cache_key}")
        return jsonify(events_cache[cache_key]['data'])

    print(f"🔄 Generando datos frescos para: {cache_key}")
    
    # 🔹 Lista de festivos
    festivos = FESTIVOS

    # 🔹 Definir el orden de los puestos
    orden_puestos = {"Administrador/a": 1, "ADM": 2, "TS": 3}
    colores_puestos = {
        "Administrador/a": "#9932CC",
        "ADM": "#B0FFB0",
        "TS": "#A3C4FF"
    }

    # 🔹 Obtener usuarios con filtros aplicados
    usuarios_query = {"visible_calendario": {"$ne": False}}
    if rol_filtro != 'todos':
        usuarios_query["puesto"] = rol_filtro
    
    usuarios = list(users_collection.find(usuarios_query))
    for usuario in usuarios:
        usuario['_id'] = str(usuario['_id'])
        usuario['nombre_completo'] = f"{usuario.get('nombre', '')} {usuario.get('apellidos', '')}".strip()
        usuario['display_name'] = get_display_name_for_user(usuario, anonymized_labels)
    usuarios_ordenados = sorted(usuarios, key=lambda u: orden_puestos.get(u.get('puesto', ''), 4))

    # 🔹 Filtrar usuarios por persona si es necesario
    if persona_nombre:
        usuarios_ordenados = [u for u in usuarios_ordenados if u['nombre_completo'] == persona_nombre]

    # 🔹 Filtrar usuarios por búsqueda si es necesario
    if busqueda:
        busqueda_lower = busqueda.lower()
        usuarios_ordenados = [
            u for u in usuarios_ordenados
            if busqueda_lower in (u['display_name'] if demo_admin_mode else u['nombre_completo']).lower()
        ]

    # 🔹 Obtener eventos solo para los usuarios filtrados
    nombres_usuarios = [u['nombre_completo'] for u in usuarios_ordenados]
    eventos_query = {"trabajador": {"$in": nombres_usuarios}} if nombres_usuarios else {}
    eventos = list(events_collection.find(eventos_query))

    # Agrupar los eventos por trabajador y por tipo
    eventos_por_trabajador = {}
    for evento in eventos:
        nombre = evento["trabajador"]  # Supongo que se guarda el nombre completo
        tipo = evento.get("tipo", "Vacaciones")
        if nombre not in eventos_por_trabajador:
            eventos_por_trabajador[nombre] = {}
        if tipo not in eventos_por_trabajador[nombre]:
            eventos_por_trabajador[nombre][tipo] = []
        eventos_por_trabajador[nombre][tipo].append((evento["fecha_inicio"], evento["fecha_fin"]))

    # 🔹 Estructura para agrupar vacaciones por trabajador
    dias_no_disponibles = {}
    for evento in eventos:
        nombre = evento["trabajador"]
        if nombre not in dias_no_disponibles:
            dias_no_disponibles[nombre] = []
        dias_no_disponibles[nombre].append((evento["fecha_inicio"], evento["fecha_fin"]))

    eventos_json = []
    contador_disponibles = {}

    # 🔹 Generar eventos desde el 1 de enero hasta el 31 de diciembre de 2025
    fecha_actual = datetime(2025, 1, 1)
    fecha_fin = datetime(2028, 1, 15)
  
    while fecha_actual <= fecha_fin:
        fecha_str = fecha_actual.strftime("%Y-%m-%d")
        dia_semana = fecha_actual.weekday()

        if fecha_str in festivos:
            eventos_json.append({
                "id": f"Festivo-{fecha_str}",
                "title": "Festivo",
                "start": fecha_str,
                "color": "#A9A9A9",
                "classNames": ["festivo-event"]
            })
            contador_disponibles[fecha_str] = 0

        elif dia_semana < 5:  # Sólo de lunes a viernes
            disponibles_en_dia = 0
            eventos_dia = []

            for usuario in usuarios_ordenados:
                nombre_completo = usuario['nombre_completo']
                display_name = usuario['display_name']
                color = colores_puestos.get(usuario["puesto"], "#D3D3D3")
                event_label = f"{usuario['puesto']} - {display_name}"

                # Verifica si hay eventos asignados para este usuario en la fecha
                user_events = eventos_por_trabajador.get(nombre_completo, {})
                evento_asignado = None
                # Verifica en orden de prioridad: Vacaciones, CADE 30, CADE 50, Mail.
                for tipo in ["Baja", "CADE 30", "CADE 50", "CADE Tardes", "Guardia CADE", "Refuerzo Cade", "Mail", "Vacaciones"]:
                    if tipo in user_events:
                        for inicio, fin in user_events[tipo]:
                            # Suponiendo que inicio y fin son strings "YYYY-MM-DD"
                            if fecha_str >= inicio and fecha_str <= fin:
                                evento_asignado = tipo
                                break
                    if evento_asignado:
                        break
                if evento_asignado:
                    if evento_asignado == "Vacaciones":
                        event_label += " (Ausente)"
                        color = "#E53935"  # Rojo brillante
                    elif evento_asignado == "Baja":
                        event_label += " (Baja)"
                        color = "#757575"  # Gris medio
                    elif evento_asignado == "CADE 30":
                        event_label += " (CADE 30)"
                        color = "#FB8C00"  # Naranja fuerte
                    elif evento_asignado == "CADE 50":
                        event_label += " (CADE 50)"
                        color = "#FDD835"  # Amarillo brillante
                    elif evento_asignado == "CADE Tardes":
                        event_label += " (CADE Tardes)"
                        color = "#F6AE2D"  # Dorado/miel vibrante
                    elif evento_asignado == "Guardia CADE":
                        event_label += " (Guardia CADE)"
                        color = "#49A275"  # Verde oscuro
                    elif evento_asignado == "Refuerzo Cade":
                        event_label += " (Refuerzo Cade)"
                        color = "#FCF2B1"  # Amarillo clarito
                    elif evento_asignado == "Mail":
                        event_label += " (Mail)"
                        color = "#8D6E63"  # Marrón rosado apagado
                else:
                    disponibles_en_dia += 1

                # Mapear el estado para el frontend (Vacaciones -> Ausente)
                estado_value = "PIAS"
                if evento_asignado:
                    if evento_asignado == "Vacaciones":
                        estado_value = "Ausente"
                    else:
                        estado_value = evento_asignado

                # 🔹 Aplicar filtro de estados en el backend
                if estados_filtro and estado_value not in estados_filtro:
                    continue

                eventos_dia.append({
                    "id": f"{usuario['_id']}-{fecha_str}",
                    "title": event_label,
                    "start": fecha_str,
                    "color": color,
                    "extendedProps": {
                        "nombre": display_name,
                        "puesto": usuario.get("puesto", ""),
                        "estado": estado_value,
                        "isDisponible": (evento_asignado is None)
                    }
                })

            eventos_dia.sort(key=lambda e: orden_puestos.get(e["title"].split(" - ")[0], 4))
            eventos_json.extend(eventos_dia)
            contador_disponibles[fecha_str] = disponibles_en_dia

        fecha_actual += timedelta(days=1)

    # 🔹 Guardar en caché
    result = {"eventos": eventos_json, "contador": contador_disponibles}
    events_cache[cache_key] = {
        'data': result,
        'timestamp': time.time()
    }
    
    return jsonify(result)

@app.route('/api/events/<event_id>', methods=['DELETE'])
@login_required
def delete_event(event_id):
    evento = events_collection.find_one({"_id": ObjectId(event_id)})
    if not evento:
        return jsonify({"message": "Evento no encontrado"}), 404

    if evento["trabajador"] != f"{current_user.nombre} {current_user.apellidos}":
        return jsonify({"message": "No puedes eliminar eventos de otros"}), 403

    events_collection.delete_one({"_id": ObjectId(event_id)})
    # Invalidar caché cuando se elimina un evento
    invalidate_cache()
    return jsonify({"message": "Evento eliminado, el usuario vuelve a estar disponible"}), 200

@app.route('/admin/asignar-estados', methods=['GET', 'POST'])
@login_required
@admin_required
def asignar_estados():
    if request.method == 'POST':
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')

        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        trabajadores = list(users_collection.find())
        
        for trabajador in trabajadores:
            estado = request.form.get(f"tipo_{trabajador['_id']}")
            if not estado:
                continue

            nombre_completo = f"{trabajador['nombre']} {trabajador['apellidos']}"

            # Iterar por cada día del rango
            current_day = fecha_inicio_dt
            while current_day <= fecha_fin_dt:
                day_str = current_day.strftime("%Y-%m-%d")

                query_filter = {
                    "trabajador": nombre_completo,
                    "fecha_inicio": day_str,
                    "fecha_fin": day_str,
                    "tipo": {"$in": ["Baja", "CADE 30", "CADE 50", "CADE Tardes", "Guardia CADE", "Refuerzo Cade", "Mail"]}
                }

                # Saltar fines de semana y festivos, pero permitir limpieza de eventos si se selecciona "normal"
                if not es_dia_habil(current_day):
                    if estado == "normal":
                        events_collection.delete_many(query_filter)
                    current_day += timedelta(days=1)
                    continue

                if estado == "normal":
                    # Intentamos eliminar los eventos especiales para este trabajador y este rango
                    events_collection.delete_many(query_filter)
                else:
                    # Primero eliminamos cualquier evento previo especial en ese rango (para evitar duplicados)
                    events_collection.delete_many(query_filter)
                    # Luego insertamos el nuevo evento
                    nuevo_evento = {
                        "trabajador": nombre_completo,
                        "fecha_inicio": day_str,
                        "fecha_fin": day_str,
                        "tipo": estado
                    }
                    events_collection.insert_one(nuevo_evento)
                    
                current_day += timedelta(days=1)
            
        invalidate_cache()  # Invalidar caché al asignar estados masivos
        return redirect(url_for('asignar_estados'))
    else:
        trabajadores = list(users_collection.find({"visible_calendario": True}))
        trabajadores = sorted(trabajadores, key=lambda x: (x['nombre'].lower(), x['apellidos'].lower()))
        demo_mode = is_demo_admin_user()
        anonymized_labels = build_anonymized_label_map() if demo_mode else None
        for trabajador in trabajadores:
            trabajador['display_name'] = get_display_name_for_user(trabajador, anonymized_labels)
            full_name = f"{trabajador.get('nombre', '')} {trabajador.get('apellidos', '')}".strip()
            trabajador['full_name'] = full_name
            search_source = trabajador['display_name'] if demo_mode else full_name
            trabajador['search_value'] = search_source.lower()
        return render_template("asignar_estados.html", trabajadores=trabajadores, hide_real_names=demo_mode)
    

@app.route('/admin/duplicados', methods=['GET'])
@login_required
@admin_required
def duplicados():
    pipeline = [
        # Excluir fines de semana antes de agrupar (independiente del tipo de campo)
        {"$addFields": {"_fechaDate": {"$toDate": "$fecha_inicio"}}},
        {"$match": {"$expr": {"$not": {"$in": [{"$isoDayOfWeek": "$_fechaDate"}, [6, 7]]}}}},
        {
            "$match": {
                "$expr": {
                    "$not": {
                        "$in": [
                            {"$dateToString": {"format": "%Y-%m-%d", "date": "$_fechaDate"}},
                            FESTIVOS_LIST
                        ]
                    }
                }
            }
        },
        {
            "$group": {
                "_id": {
                    "trabajador": "$trabajador",
                    "fecha_inicio": "$fecha_inicio",
                    "fecha_fin": "$fecha_fin"
                },
                "tipos": {"$addToSet": "$tipo"},
                "count": {"$sum": 1},
                "ids": {"$push": "$_id"}
            }
        },
        {
            "$match": {
                "$expr": {"$gt": [{"$size": "$tipos"}, 1]}
            }
        },
        {"$sort": {"_id.fecha_inicio": 1}},
        {"$project": {"_fechaDate": 0}}
    ]
    duplicados = list(events_collection.aggregate(pipeline))
    demo_mode = is_demo_admin_user()
    fullname_labels = None
    if demo_mode:
        _, fullname_labels = build_anonymized_label_map(visible_only=False, include_fullname_map=True)
    for dup in duplicados:
        real_name = dup["_id"].get("trabajador", "")
        if fullname_labels:
            dup["display_trabajador"] = fullname_labels.get(real_name, real_name)
        else:
            dup["display_trabajador"] = real_name
    return render_template("duplicados.html", duplicados=duplicados, hide_real_names=demo_mode)


def _parse_metrics_date(value):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None


def resolver_rango_metricas(fecha_inicio, fecha_fin):
    """Normalize metrics date range and return business days count."""
    today = datetime.today().date()

    start_date = _parse_metrics_date(fecha_inicio)
    end_date = _parse_metrics_date(fecha_fin)

    if start_date is None and end_date is None:
        start_date = today.replace(day=1)
        end_day = monthrange(today.year, today.month)[1]
        end_date = date(today.year, today.month, end_day)
    elif start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date
    elif start_date and not end_date:
        end_date = start_date
    elif end_date and not start_date:
        start_date = end_date

    start_str = start_date.strftime("%Y-%m-%d") if start_date else None
    end_str = end_date.strftime("%Y-%m-%d") if end_date else None

    if start_date and end_date:
        dias_periodo = contar_dias_habiles_en_rango(start_str, end_str)
    else:
        dias_periodo = 0

    return start_str, end_str, dias_periodo


def aplicar_pias(metricas, dias_periodo):
    total_pias = 0
    dias_periodo_int = int(dias_periodo or 0)
    for datos in metricas.values():
        dias_ocupados = 0
        for clave, valor in datos.items():
            if clave in ('PIAS', 'CADE Total'):
                continue
            try:
                dias_ocupados += int(valor)
            except (TypeError, ValueError):
                continue
        pias_valor = dias_periodo_int - dias_ocupados
        if pias_valor < 0:
            pias_valor = 0
        datos['PIAS'] = pias_valor
        total_pias += pias_valor
    return total_pias


def calcular_metricas_por_usuario(fecha_inicio=None, fecha_fin=None, puesto=None):
    """Devuelve un diccionario con el conteo de eventos por tipo para cada trabajador
    y el ranking de los 5 con más registros por tipo. Se pueden filtrar los
    eventos por rango de fechas."""

    match_stage = {}
    if fecha_inicio and fecha_fin:
        match_stage = {"fecha_inicio": {"$gte": fecha_inicio, "$lte": fecha_fin}}
    elif fecha_inicio:
        match_stage = {"fecha_inicio": {"$gte": fecha_inicio}}
    elif fecha_fin:
        match_stage = {"fecha_inicio": {"$lte": fecha_fin}}

    trabajadores_query = {"visible_calendario": {"$ne": False}}
    pipeline = []
    if puesto:
        puesto_db = "Administrador/a" if puesto.lower() in ["admin", "administrador", "administrador/a"] else puesto
        trabajadores_query["puesto"] = puesto_db

    trabajadores_base = [
        f"{u['nombre']} {u['apellidos']}".strip()
        for u in users_collection.find(trabajadores_query, {"nombre": 1, "apellidos": 1})
    ]

    if puesto:
        if trabajadores_base:
            pipeline.append({"$match": {"trabajador": {"$in": trabajadores_base}}})
        else:
            pipeline.append({"$match": {"_id": {"$exists": False}}})
    if match_stage:
        pipeline.append({"$match": match_stage})
    
    pipeline.append({
        "$addFields": {
            "fecha_date": {
                "$switch": {
                    "branches": [
                        {
                            "case": {"$eq": [{"$type": "$fecha_inicio"}, "date"]},
                            "then": "$fecha_inicio"
                        },
                        {
                            "case": {"$eq": [{"$type": "$fecha_inicio"}, "string"]},
                            "then": {"$dateFromString": {"dateString": "$fecha_inicio"}}
                        }
                    ],
                    "default": {"$toDate": "$fecha_inicio"}
                }
            }
        }
    })
    pipeline.append({"$match": {"fecha_date": {"$ne": None}}})
    pipeline.append({
        "$match": {
            "$expr": {
                "$and": [
                    {
                        "$not": [
                            {
                                "$in": [
                                    {"$dateToString": {"format": "%Y-%m-%d", "date": "$fecha_date"}},
                                    FESTIVOS_LIST
                                ]
                            }
                        ]
                    },
                    {"$ne": [{"$dayOfWeek": "$fecha_date"}, 1]},
                    {"$ne": [{"$dayOfWeek": "$fecha_date"}, 7]}
                ]
            }
        }
    })

    pipeline.append({
        "$group": {
            "_id": {
                "trabajador": "$trabajador",
                "tipo": "$tipo",
                "fecha": {"$dateToString": {"format": "%Y-%m-%d", "date": "$fecha_date"}}
            }
        }
    })
    pipeline.append({
        "$group": {
            "_id": {"trabajador": "$_id.trabajador", "tipo": "$_id.tipo"},
            "count": {"$sum": 1}
        }
    })
    resultado = list(events_collection.aggregate(pipeline))

    estados_base = [estado for estado in events_collection.distinct("tipo") if estado and estado != "PIAS"]

    metricas = {}
    estados_periodo = set()
    top5_por_tipo = {}

    for item in resultado:
        trabajador = item["_id"]["trabajador"]
        tipo = item["_id"].get("tipo", "Desconocido")
        estados_periodo.add(tipo)

        if trabajador not in metricas:
            metricas[trabajador] = {}
        metricas[trabajador][tipo] = item["count"]

        if tipo not in top5_por_tipo:
            top5_por_tipo[tipo] = []
        top5_por_tipo[tipo].append((trabajador, item["count"]))

    for trabajador in trabajadores_base:
        metricas.setdefault(trabajador, {})

    # Ordenar alfabéticamente los trabajadores
    metricas = dict(sorted(metricas.items(), key=lambda x: x[0]))



    estados_final = sorted(set(estados_base) | estados_periodo, key=lambda s: str(s).lower())

    for datos in metricas.values():
        for estado in estados_final:
            datos.setdefault(estado, 0)

    cade_estados = ("CADE 30", "CADE 50", "CADE Tardes")
    cade_label = "CADE Total"

    for datos in metricas.values():
        sum_cade = 0
        for estado in cade_estados:
            try:
                sum_cade += int(datos.get(estado, 0))
            except (TypeError, ValueError):
                continue
        datos[cade_label] = sum_cade

    if cade_label not in estados_final:
        estados_list = list(estados_final)
        insert_pos = -1
        for estado in cade_estados:
            if estado in estados_list:
                idx = estados_list.index(estado)
                if idx > insert_pos:
                    insert_pos = idx
        if insert_pos >= 0:
            estados_list.insert(insert_pos + 1, cade_label)
        else:
            estados_list.append(cade_label)
        estados_final = estados_list
    else:
        estados_final = list(estados_final)

    # Seleccionar top 5 por tipo
    for tipo, lista in top5_por_tipo.items():
        lista.sort(key=lambda x: x[1], reverse=True)
        top5_por_tipo[tipo] = lista[:5]

    return metricas, estados_final, top5_por_tipo


def anonymize_metric_results(metricas, top5_por_tipo, fullname_labels):
    """Devuelve copias de metricas/top5 usando etiquetas anonimizadas."""
    if not fullname_labels:
        return metricas, top5_por_tipo

    top5_por_tipo = top5_por_tipo or {}
    masked_metricas = {}
    for nombre, datos in metricas.items():
        label = fullname_labels.get(nombre, nombre)
        masked_metricas[label] = dict(datos)
    masked_metricas = dict(sorted(masked_metricas.items(), key=lambda x: x[0]))

    masked_top5 = {}
    for tipo, lista in top5_por_tipo.items():
        masked_top5[tipo] = [
            (fullname_labels.get(nombre, nombre), count)
            for nombre, count in lista
        ]

    return masked_metricas, masked_top5


@app.route('/dashboard-metrics')
@login_required
@admin_required
def dashboard_metrics():
    fecha_inicio_raw = request.args.get('fecha_inicio')
    fecha_fin_raw = request.args.get('fecha_fin')
    puesto = request.args.get('puesto')
    fecha_inicio, fecha_fin, dias_periodo = resolver_rango_metricas(fecha_inicio_raw, fecha_fin_raw)
    metricas, estados, top5_por_tipo = calcular_metricas_por_usuario(fecha_inicio, fecha_fin, puesto)
    demo_mode = is_demo_admin_user()
    if demo_mode:
        _, fullname_labels = build_anonymized_label_map(visible_only=False, include_fullname_map=True)
        metricas, top5_por_tipo = anonymize_metric_results(metricas, top5_por_tipo, fullname_labels)

    total_pias = aplicar_pias(metricas, dias_periodo)
    return render_template(
        'dashboard_metrics.html',
        metricas=metricas,
        estados=estados,
        top5_por_tipo=top5_por_tipo,
        dias_periodo=dias_periodo,
        total_pias=total_pias,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        puesto=puesto,
        hide_real_names=demo_mode
    )


@app.route('/dashboard-metrics/export', methods=['GET'])
@login_required
@admin_required
def dashboard_metrics_export():
    """Exporta la tabla de métricas a Excel (.xlsx) o CSV (fallback)."""
    use_xlsx = True
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except Exception:
        use_xlsx = False

    fecha_inicio_raw = request.args.get('fecha_inicio')
    fecha_fin_raw = request.args.get('fecha_fin')
    puesto = request.args.get('puesto')

    fecha_inicio, fecha_fin, dias_periodo = resolver_rango_metricas(fecha_inicio_raw, fecha_fin_raw)
    metricas, estados, _ = calcular_metricas_por_usuario(fecha_inicio, fecha_fin, puesto)
    demo_mode = is_demo_admin_user()
    if demo_mode:
        _, fullname_labels = build_anonymized_label_map(visible_only=False, include_fullname_map=True)
        metricas, _ = anonymize_metric_results(metricas, None, fullname_labels)
    aplicar_pias(metricas, dias_periodo)
    headers = ["Trabajador", "PIAS"] + list(estados)

    filename_parts = ["metricas"]
    if fecha_inicio:
        filename_parts.append(fecha_inicio)
    if fecha_fin:
        filename_parts.append(fecha_fin)
    if puesto:
        filename_parts.append(("admin" if puesto == "Administrador/a" else puesto).lower())

    if use_xlsx:
        wb = Workbook()
        ws = wb.active
        ws.title = "Métricas"

        ws.append(headers)
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
        center = Alignment(horizontal="center")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for trabajador, datos in metricas.items():
            row = [trabajador, int(datos.get('PIAS', 0))] + [int(datos.get(estado, 0)) for estado in estados]
            ws.append(row)

        ws.column_dimensions['A'].width = 40
        for i in range(2, len(headers) + 1):
            letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[letter].width = 14

        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c > 1:
                    cell.alignment = center

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = "_".join(filename_parts) + ".xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        # Fallback a CSV para compatibilidad sin dependencias
        import csv
        from io import StringIO
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        for trabajador, datos in metricas.items():
            row = [trabajador, int(datos.get('PIAS', 0))] + [int(datos.get(estado, 0)) for estado in estados]
            writer.writerow(row)
        output = BytesIO(sio.getvalue().encode('utf-8-sig'))
        filename = "_".join(filename_parts) + ".csv"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )


# Configuramos asistente de IA

import openai
import pinecone
from pinecone import Pinecone
from openai import OpenAI
from procesar_pdfs import guardar_pdf_en_pinecone, extraer_texto_pdf


OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# 🔹 Configurar Pinecone
PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")
PINECONE_ENVIRONMENT = os.getenv("PINECONE_ENVIRONMENT")
PINECONE_INDEX_NAME = os.getenv("PINECONE_INDEX_NAME")
# Inicializar Pinecone correctamente con la nueva sintaxis
pc = Pinecone(api_key=PINECONE_API_KEY)

# 🔹 Inicializar cliente de OpenAI
client = OpenAI()

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# Verificar si el índice existe antes de usarlo
if PINECONE_INDEX_NAME not in pc.list_indexes().names():
    print(f"❌ El índice '{PINECONE_INDEX_NAME}' no existe en Pinecone.")
else:
    index = pc.Index(PINECONE_INDEX_NAME)


@app.route('/ai-response', methods=['POST'])
@login_required
def ai_response():
    data = request.get_json()
    user_message = data.get("message")  # Extraer el mensaje del usuario

    if not user_message:
        return jsonify({"error": "Mensaje vacío"}), 400

    # 🔹 Obtener los resultados de Pinecone (lista de tuplas: (texto, documento, página))
    resultados = buscar_en_pinecone(user_message)

    # Separar el texto relevante de cada fragmento
    context_texts = [r[0] for r in resultados]
    contexto_str = "\n".join(context_texts) if context_texts else "No se encontró información relevante."

    # Construir la lista de fuentes, incluyendo el documento y la página
    # Se crea un set para evitar duplicados
    fuentes_set = {(r[1], r[2]) for r in resultados if r[1] != "Desconocido"}
    doc_str = ", ".join([f"{name} (página {page})" for name, page in fuentes_set]) if fuentes_set else "No se encontraron documentos relevantes."

    # 🔹 Recuperar el historial de conversación (últimos N mensajes) si lo usas
    historial = list(historial_collection.find({"usuario": current_user.usuario}).sort("timestamp", -1).limit(5))
    historial_str = "\n".join([f"Usuario: {h['mensaje']}\nAsistente: {h['respuesta']}" for h in historial])

    # 🔹 Generar respuesta con OpenAI
    prompt = f"""
    Contexto relevante recuperado:
    {contexto_str}

    Historial de conversación:
    {historial_str}

    Usuario: {user_message}
    Asistente:
    """

    response = client.chat.completions.create(
        model="o3-mini",
        messages=[
            {"role": "system", "content": "Actúa como un asistente experto en la gestión de expedientes de dependencia que ayuda a los trabajadores de la sección PIAS del Instituto Aragonés de Servicios Sociales (IASS) a buscar información. Tu ámbito de conocimiento se centra exclusivamente en los procesos y la herramienta informática utilizada en esta sección. Cuando un trabajador te consulte, responde de forma concisa y directa a su pregunta, proporcionando la información o la guía necesaria para resolver su duda o avanzar en el proceso. Prioriza la claridad y la utilidad práctica en tus respuestas. Evita dar información que no esté directamente relacionada con la gestión de expedientes de dependencia en la herramienta informática del IASS. Recuerda que tu objetivo principal es ser una herramienta de apoyo eficaz y eficiente para los empleados."},
            {"role": "user", "content": prompt}
        ]
    )

    respuesta_final = response.choices[0].message.content

    # 🔹 Guardar en MongoDB el historial de conversación
    guardar_historial(current_user.usuario, user_message, respuesta_final)

    # Enviar también las fuentes (doc_str) en la respuesta
    return jsonify({
        "response": respuesta_final,
        "sources": doc_str
    })



def guardar_texto_en_pinecone(texto, metadata={}):
    """Convierte un texto en embeddings y lo almacena en Pinecone."""
    
    # 🔹 Generar embeddings con OpenAI
    response = openai.embeddings.create(
        model="text-embedding-ada-002",
        input=[texto]  # Debe ser una lista
    )
    
    embedding = response["data"][0]["embedding"]

    # 🔹 Guardar en Pinecone
    id_vector = str(uuid.uuid4())  # Generar un ID único para el vector
    index.upsert([(id_vector, embedding, metadata)])
    
    return id_vector

def buscar_en_pinecone(texto, documento=None):
    response = openai.embeddings.create(
        model="text-embedding-ada-002",
        input=texto
    )
    embedding = response.data[0].embedding

    filtro = {"documento": {"$eq": documento}} if documento else {}

    resultados = index.query(vector=embedding, top_k=5, include_metadata=True, filter=filtro)

    # Devolver una lista de tuplas (texto_del_fragmento, nombre_del_documento, pagina)
    return [
        (
            res["metadata"].get("texto", ""),
            res["metadata"].get("documento", "Desconocido"),
            res["metadata"].get("pagina", "N/D")
        )
        for res in resultados.get("matches", [])
    ]

def guardar_historial(usuario, mensaje, respuesta):
    """Guarda una conversación en MongoDB."""
    historial_collection.insert_one({
        "usuario": usuario,
        "mensaje": mensaje,
        "respuesta": respuesta,
        "timestamp": datetime.now(timezone.utc)
    })

"""@app.route("/upload", methods=["GET", "POST"])
@login_required
@admin_required
def upload_pdf():
    if request.method == "POST":
        if "file" not in request.files:
            return "❌ No se subió ningún archivo", 400
        
        file = request.files["file"]
        if file.filename == "":
            return "❌ No se seleccionó ningún archivo", 400

        # 🔹 Guardar archivo en la carpeta 'uploads'
        pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(pdf_path)

        # 🔹 Procesar el PDF y guardarlo en Pinecone
        resultado = guardar_pdf_en_pinecone(pdf_path)
        
        return jsonify({"message": resultado})

    return render_template("upload.html")  # Página para subir archivos
    """

"""@app.route("/search", methods=["GET", "POST"])
def search():
    documentos = index.describe_index_stats()["namespaces"].keys()  # 🔹 Obtener documentos en Pinecone
    resultados = []

    if request.method == "POST":
        query = request.form.get("query")
        documento = request.form.get("documento") or None  # None significa "todos los documentos"

        resultados = buscar_en_pinecone(query, documento)

    return render_template("search.html", documentos=documentos, resultados=resultados)

"""

# 🔹 Ruta para subir archivos desde la web
"""@app.route("/subir-pdf", methods=["GET", "POST"])
@login_required
@admin_required
def subir_pdf():
    if request.method == "POST":
        if "archivo" not in request.files:
            return jsonify({"error": "No se ha seleccionado ningún archivo."}), 400

        archivo = request.files["archivo"]

        if archivo.filename == "":
            return jsonify({"error": "El archivo no tiene nombre válido."}), 400

        if archivo and archivo.filename.endswith(".pdf"):
            archivo_path = os.path.join(app.config["UPLOAD_FOLDER"], archivo.filename)
            archivo.save(archivo_path)

            # Guardar el PDF en Pinecone
            resultado = guardar_pdf_en_pinecone(archivo_path, archivo.filename)
            
            # Opcional: puedes almacenar un mensaje de éxito en flash para mostrarlo en el dashboard
            from flask import flash
            flash(resultado)
            
            return redirect(url_for('dashboard'))

        return jsonify({"error": "Formato no válido. Solo se permiten archivos PDF."}), 400

    return render_template("subir_pdf.html")
"""
import os
import re

def custom_filename(filename):
    """
    Limpia el nombre del archivo permitiendo espacios, puntos y guiones.
    Elimina caracteres no deseados, pero conserva los espacios.
    """
    filename = os.path.basename(filename)
    # Permite letras, dígitos, espacios, guiones y puntos.
    filename = re.sub(r'[^\w\s\.-]', '', filename)
    # Reemplaza múltiples espacios por uno solo y quita espacios al inicio/final.
    filename = re.sub(r'\s+', ' ', filename).strip()
    return filename

@app.route("/subir-pdf", methods=["GET", "POST"])
@login_required
@admin_required
def subir_pdf():
    if request.method == "POST":
        if "archivo" not in request.files:
            return jsonify({"error": "No se ha seleccionado ningún archivo."}), 400

        archivo = request.files["archivo"]
        if archivo.filename == "":
            return jsonify({"error": "El archivo no tiene nombre válido."}), 400

        if archivo and archivo.filename.lower().endswith(".pdf"):
            # Usa la función personalizada para normalizar el nombre y conservar espacios.
            filename = custom_filename(archivo.filename)
            temp_path = os.path.join("uploads", filename)
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            # Guarda el archivo temporalmente usando un bloque with para asegurarte de cerrarlo.
            with open(temp_path, 'wb') as f:
                f.write(archivo.read())
            
            try:
                # Sube el archivo a S3.
                s3_client.upload_file(
                    temp_path,
                    os.environ.get("AWS_S3_BUCKET"),
                    f"uploads/{filename}",
                    ExtraArgs={"ContentType": archivo.content_type}
                )
                file_url = f"https://{os.environ.get('AWS_S3_BUCKET')}.s3.{os.environ.get('AWS_S3_REGION')}.amazonaws.com/uploads/{filename}"
                
                # Procesa el PDF y almacena en Pinecone.
                resultado = guardar_pdf_en_pinecone(temp_path, filename, index)
        
                
                return redirect(url_for('dashboard'))
            except Exception as e:
                print("Error al subir a S3 o procesar en Pinecone:", e, flush=True)
                return jsonify({"error": "Error al subir el archivo."}), 500

        return jsonify({"error": "Formato no válido. Solo se permiten archivos PDF."}), 400

    return render_template("subir_pdf.html")


@app.route("/documentos_subidos_s3", methods=["GET"])
@login_required  # O el nivel de restricción que prefieras
def documentos_subidos_s3():
    try:
        response = s3_client.list_objects_v2(
            Bucket=AWS_S3_BUCKET,
            Prefix="uploads/"
        )
        # response.get("Contents") es una lista de objetos
        files = []
        for obj in response.get("Contents", []):
            # Extraemos el nombre del archivo
            key = obj["Key"]
            # Construimos la URL del archivo
            file_url = f"https://{AWS_S3_BUCKET}.s3.{AWS_S3_REGION}.amazonaws.com/{key}"
            files.append({"name": key.split("/")[-1], "url": file_url})
    except Exception as e:
        print("Error al listar archivos en S3:", e)
        files = []
    
    return render_template("documentos_subidos.html", files=files)

@app.route("/eliminar_documento", methods=["POST"])
@login_required
@admin_required
def eliminar_documento():
    filename = request.form.get("filename")
    
    if not filename:
        flash("No se especificó ningún archivo para eliminar", "error")
        return redirect(url_for("documentos_subidos_s3"))
    
    try:
        # 1. Eliminar el archivo de S3
        s3_object_key = f"uploads/{filename}"
        s3_client.delete_object(
            Bucket=AWS_S3_BUCKET,
            Key=s3_object_key
        )
        
        # 2. Eliminar los vectores asociados de Pinecone
        try:
            # Obtener todos los vectores con el mismo nombre de documento
            fetch_response = index.query(
                vector=[0] * 1536,  # Vector de ceros de la dimensión correcta
                filter={"documento": {"$eq": filename}},
                top_k=10000,  # Aumentar para manejar más vectores
                include_metadata=True
            )
            
            matches = fetch_response.matches
            if matches:
                ids_to_delete = [match.id for match in matches]
                # Eliminar en lotes de 1000 para evitar límites potenciales
                for i in range(0, len(ids_to_delete), 1000):
                    batch = ids_to_delete[i:i + 1000]
                    index.delete(ids=batch)
                print(f"Se eliminaron {len(ids_to_delete)} vectores de Pinecone para '{filename}'")
            else:
                print(f"No se encontraron vectores en Pinecone para '{filename}'")
                
        except Exception as e:
            print(f"Error al eliminar vectores de Pinecone: {str(e)}")
            flash(f"Archivo eliminado de S3, pero hubo un error al eliminar de Pinecone: {str(e)}", "warning")
            return redirect(url_for("documentos_subidos_s3"))
        
        flash(f"El documento '{filename}' ha sido eliminado correctamente de S3 y Pinecone", "success")
        return redirect(url_for("documentos_subidos_s3"))
        
    except Exception as e:
        flash(f"Error al eliminar el documento: {str(e)}", "error")
        return redirect(url_for("documentos_subidos_s3"))
    

@app.route('/informe_uso_ia', methods=['GET'])
@login_required
@admin_required
def informe_uso_ia():
    # Obtener fecha de la URL
    fecha_str = request.args.get("fecha")
    fecha_dt = None
    if fecha_str:
        try:
            fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        except ValueError:
            fecha_dt = None

    # Cargar FAQs generadas desde JSON
    fecha_faqs = None
    fecha_origen = None
    faqs = []
    if os.path.exists(ruta_faqs):
        with open(ruta_faqs, "r", encoding="utf-8") as f:
            datos_faq = json.load(f)
            faqs = datos_faq.get("faqs", [])
            # Ordenar FAQs por frecuencia de mayor a menor
            faqs.sort(key=lambda x: x.get("frecuencia", 0), reverse=True)            
            fecha_faqs = datos_faq.get("fecha_generacion")
            fecha_origen = datos_faq.get("fecha_origen")

    # Calcular consultas por usuario a partir de MongoDB
    filtro = {}
    if fecha_dt:
        filtro["timestamp"] = {"$gte": fecha_dt}

    mensajes = list(historial_collection.find(filtro))

    consultas_por_usuario = {}
    for m in mensajes:
        usuario = m.get("usuario", "Desconocido")
        consultas_por_usuario[usuario] = consultas_por_usuario.get(usuario, 0) + 1

    # Ordenar las consultas por usuario de mayor a menor
    consultas_por_usuario = sorted(consultas_por_usuario.items(), key=lambda x: x[1], reverse=True)

    return render_template(
        "informe_uso_ia.html",
        faqs=faqs,
        consultas_por_usuario=consultas_por_usuario,
        fecha_faqs=fecha_faqs,
        fecha_origen=fecha_origen
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=True, use_reloader=False, use_debugger=False)
